import openpyxl

#syntax to use before write or read on excel sheet
file_path = r"C:\Users\HP\PycharmProjects\selenium\write.xlsx"
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
rows = worksheet.max_row
cols = worksheet.max_column

# #to write on excel sheet
# for rows in range(1,5+1): #selecting number of rows
#     for cols in range(1,6+1):  #selecting number of columns
#        worksheet.cell(row=rows, column=cols).value='selenium' #syntax to write on excel
# workbook.save(file_path)


# to read from excel
for rows in range(1, rows+1):
    for cols in range(1, cols+1):
        print(worksheet.cell(row=rows, column=cols).value)
    print()
