import openpyxl
file_path = r"C:\Users\HP\Desktop\sample.xlsx"
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
rows = worksheet.max_row
cols = worksheet.max_column
'''
values = []
for row in range(1, rows+1):
    for col in range(1, cols+1):
        a = worksheet.cell(row=row, column=col).value
        values+= [(a)]
    print()
print(values)
'''
values = [worksheet.cell(row=row, column=col).value for row in range(1, rows+1) for col in range(1, cols+1)]
names, password = [], []
print(values)
# '''
for i in range(2,len(values)):
    if i%2 == 0:
        names.append(values[i])
    else:
        password.append(values[i])
for i, j in zip(names, password):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    import time
    driver = webdriver.Chrome(r"C:\Users\HP\PycharmProjects\selenium\drivers\chromedriver.exe")
    url = r"https://www.facebook.com/login/"
    driver.get(url)
    time.sleep(2)
    driver.find_element(By.ID, "email").send_keys(i)
    time.sleep(2)
    driver.find_element(By.ID, "pass").send_keys(j)
    time.sleep(2)
    driver.close()